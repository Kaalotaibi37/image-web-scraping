from openpyxl import Workbook, load_workbook
import PySimpleGUI as sg
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import time


def download_tables():

    # Read in target excel spreadsheet
    wb_url = load_workbook("Sample.xlsx")
    first_sheet = wb_url.sheetnames[0]
    ws_url = wb_url[first_sheet]

    # Set the target columns
    code_col = "A"
    url_col = "B"
    urls = []
    barcodes = []

    # Starting from 2 to avoid header row
    for row in range(2, ws_url.max_row + 1):
        col = code_col
        cell_name = "{}{}".format(col, row)
        barcodes.append(ws_url[cell_name].value)
        col = url_col
        cell_name = "{}{}".format(col, row)
        urls.append(ws_url[cell_name].value)

    # Start selenium
    options = Options()
    options.add_argument("--headless")
    chromedriver = "chromedriver.exe"
         
    driver = webdriver.Chrome(executable_path=chromedriver, options=options)

    for urlIdx in range(len(urls)):

        # For every 50 urls, stop and rest for 5 seconds
        if urlIdx % 50 == 0:
            time.sleep(5)
            
        url = urls[urlIdx]
        barcode = barcodes[urlIdx]
        
        try:
            driver.get(url)

            # Find nutrition table image and download if it exists
            nutrition_facts = driver.find_elements_by_xpath("//figure[@id='image_box_nutrition']")
            nutrition_fact = nutrition_facts[1]

            nutrition_image = nutrition_fact.find_element_by_class_name("hide-for-xlarge-up")

            img_src = nutrition_image.get_attribute("src").replace(".400.", ".full.").replace(".200.", ".full.")
            driver.get(img_src)
            
            driver.get_screenshot_as_file(str(barcode) + ".png")
        
            print(f"barcode {barcode} download success")
        except:
            print(f"barcode {barcode} no download")
            continue

    driver.quit()
    return 0

if __name__ == "__main__":
    download_tables()
